import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, Alignment

from io import StringIO

from django.shortcuts import render
from django.http import HttpResponse

from .models import Stevne, ØVELSER, KLASSER

# Create your views here.

klassefarger = {
    'A'           :'#0070C0',
    'B'           :'#00B050',
    'C'           :'#974807',
    'D'           :'#000000',
    'K1'          :'#808080',
    'K2'          :'#808080',
    'J'           :'#808080',
    'Jm'          :'#808080',
    'Jk'          :'#808080',
    'U'           :'#808080',
    'U16'         :'#808080',
    'U14'         :'#808080',
    'U12'         :'#808080',
    'Åpen rekrutt':'#808080',
    'V50'         :'#808080',
    'V55'         :'#808080',
    'V60'         :'#808080',
    'V65'         :'#808080',
    'V70'         :'#808080',
    'SH1'         :'#808080'
}

def index(request):
    context = {'stevner':Stevne.objects.all()}
    return render(request, 'resultater/index.html', context)

def stevne_html(request, stevnenr):
    try:
        s = Stevne.objects.get(navn=stevnenr)
    except:
        return HttpResponse("Stevnet {} finnes ikke".format(stevnenr))

    figurer = (s.figurer1,s.figurer2,s.figurer3,s.figurer4,s.figurer5,s.figurer6,s.figurer7,s.figurer8,s.figurer9,s.figurer10)[:s.standplasser]

    maxsum_fgmr = sum(6 + min(f,6) for f in figurer)
    maxsum_spes = sum(5 + min(f,5) for f in figurer)

    context = {
        'navn':s.navn,
        'stevnetype':s.stype,
        'arrangør':s.arrangør,
        'dato':str(s.dato),
        'sted':s.sted,
        'stevneleder':s.stevneleder,
        'premiering':s.premiering,
        'figurer':figurer,
        'maxsum_fgmr':maxsum_fgmr,
        'maxsum_spes':maxsum_spes,
        'regneark':s.get_xlsx_url()
    }

    resultatliste = lag_resultatliste(s)

    context['resultater'] = resultatliste

    return render(request, 'resultater/stevne.html', context)

def stevne_xlsx(request, stevnenr):
    try:
        s = Stevne.objects.get(navn=stevnenr)
    except:
        return HttpResponse("Stevnet {} finnes ikke".format(stevnenr))

    rl = lag_resultatliste(s)
    rl.sort(key=lambda x: x['rekkefølge'])

    figurer = (s.figurer1,s.figurer2,s.figurer3,s.figurer4,s.figurer5,s.figurer6,s.figurer7,s.figurer8,s.figurer9,s.figurer10)[:s.standplasser]
    maxsum_fgmr = sum(6 + min(f,6) for f in figurer)
    maxsum_spes = sum(5 + min(f,5) for f in figurer)

    wb = Workbook()
    ws = wb.active

    ws['B01'] = 'Resultatliste'
    ws['B01'].font = Font(size=16, bold=True)
    ws['B03'] = 'Stevne';                    ws['C03'] = 'Åpent stevne'
    ws['B04'] = 'Program';                   ws['C04'] = 'TBD'
    ws['B05'] = 'Arrangør';                  ws['C05'] = s.arrangør
    ws['B06'] = 'Dato';                      ws['C06'] = str(s.dato)
    ws['B07'] = 'Sted';                      ws['C07'] = s.sted
    ws['B08'] = 'Stevnenummer';              ws['C08'] = s.navn
    ws['B09'] = 'Stevneleder';               ws['C09'] = s.stevneleder
    ws['B10'] = 'Antall deltakere';          ws['C10'] = 'TBD'
    ws['B11'] = 'Premiering';                ws['C11'] = s.premiering
    ws['B12'] = 'Oppnåelig poengsum (spes)'; ws['C12'] = '{} ({})'.format(maxsum_fgmr, maxsum_spes)
    ws['B13'] = 'Maks antall innertreff';    ws['C13'] = 'Tja'

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 23
    ws.column_dimensions['C'].width = 15
    for kol in 'DEFGHIJKLMNOPQRSTUVW':
        ws.column_dimensions[kol].width = 4

    rad = 15
    for ød in rl:
        ws.cell(row=rad, column=2, value=ød['øvelse']).font=Font(bold=True)
        rad += 1
        for kd in ød['klasser']:
            ws.cell(row=rad, column=2, value=kd['klasse']).font=Font(bold=True)
            rad += 1
            c=ws.cell(row=rad, column=1, value='Nr.');c.font=Font(bold=True);c.alignment = Alignment(horizontal='center')
            c=ws.cell(row=rad, column=2, value='Navn');c.font=Font(bold=True)
            c=ws.cell(row=rad, column=3, value='Klubb');c.font=Font(bold=True)
            for stpl in range(s.standplasser):
                c=ws.cell(row=rad, column=4+stpl, value=stpl+1);c.font=Font(bold=True);c.alignment = Alignment(horizontal='center')
            c=ws.cell(row=rad, column=5+stpl, value='SUM');c.font=Font(bold=True,size=9);c.alignment = Alignment(horizontal='center')
            c=ws.cell(row=rad, column=6+stpl, value='(*)');c.font=Font(bold=True);c.alignment = Alignment(horizontal='center')
            rad += 1
            for nr, sk in enumerate(kd['skyttere'], 1):
                ws.cell(row=rad, column=1, value=nr).alignment = Alignment(horizontal='center')
                ws.cell(row=rad, column=2, value=sk['navn'])
                ws.cell(row=rad, column=3, value=sk['klubb'])
                for stpl, p in enumerate(sk['poeng']):
                    ws.cell(row=rad, column=4+stpl, value=p).alignment = Alignment(horizontal='center')
                c=ws.cell(row=rad, column=5+stpl, value=sk['poengsum']);c.font = Font(bold=True);c.alignment = Alignment(horizontal='center')
                c=ws.cell(row=rad, column=6+stpl, value=sk['soner']);c.font = Font(italic=True);c.alignment = Alignment(horizontal='center')
                rad += 1
            rad += 1

    return HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def lag_resultatliste(s, lag_plot=False):
    resultatmap={}
    for start in s.starter.all():
        ss = {}

        ss['navn']       = start.navn
        ss['klubb']      = start.klubb
        ss['poeng']      =     (start.poeng1,start.poeng2,start.poeng3,start.poeng4,start.poeng5,start.poeng6,start.poeng7,start.poeng8,start.poeng9,start.poeng10)[:s.standplasser]
        ss['soner']      = sum((start.soner1,start.soner2,start.soner3,start.soner4,start.soner5,start.soner6,start.soner7,start.soner8,start.soner9,start.soner10)[:s.standplasser])
        ss['poengsum']   = sum(ss['poeng'])
        ss['rekkefølge'] = ss['poengsum'] * 60 + ss['soner']

        if start.øvelse not in resultatmap:
            resultatmap[start.øvelse] = {
                'rekkefølge' : [ø[1] for ø in ØVELSER].index(start.øvelse),
                'klasser'    : {}
            }

        if start.klasse not in resultatmap[start.øvelse]['klasser']:
            resultatmap[start.øvelse]['klasser'][start.klasse] = {
                'rekkefølge' : [k[1] for k in KLASSER].index(start.klasse),
                'skyttere'   : []
            }

        resultatmap[start.øvelse]['klasser'][start.klasse]['skyttere'].append(ss)

    # konverter dicts til lister av dicts for å kunne bruke dictsort
    # generer plot for hver øvelse
    resultatliste = []
    for øvelse,øvelsedata in resultatmap.items():
        øvelsedata['øvelse'] = øvelse
        øvelsedata['klasser'] = [dict(kd,klasse=kn) for kn,kd in øvelsedata['klasser'].items()]

        øvelsedata['klasser'].sort(key=lambda x: x['rekkefølge'])

        poengliste  = [[s['poengsum'] for s in k['skyttere']]for k in øvelsedata['klasser']]
        klasseliste = [k['klasse'] for k in øvelsedata['klasser']]
        fargeliste  = [klassefarger[k] for k in klasseliste]
        if lag_plot:
            plt.clf()
            plt.figure(figsize=(6,4))
            plt.hist(poengliste, bins=maxsum_fgmr+1, range=(0,maxsum_fgmr), color=fargeliste, stacked=True)
            plt.legend(klasseliste, framealpha=0.5, loc='upper left')
            svg = StringIO()
            plt.savefig(svg, format='svg')
            øvelsedata['plot'] = svg.getvalue()

        resultatliste.append(øvelsedata)

    return resultatliste
